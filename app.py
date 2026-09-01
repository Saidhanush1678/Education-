import os
import re
import sqlite3
import uuid
from datetime import datetime
from functools import wraps

import qrcode

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    send_file,
    send_from_directory,
    session,
    abort
)

from werkzeug.utils import secure_filename

from openpyxl import (
    Workbook,
    load_workbook
)

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)

from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import (
    getSampleStyleSheet,
    ParagraphStyle
)
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.units import mm

from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image as RLImage
)


# =========================================================
# FLASK APPLICATION
# =========================================================

app = Flask(__name__)


# =========================================================
# SECRET KEY
# =========================================================

app.secret_key = os.environ.get(
    "SECRET_KEY",
    "change-this-secret-key-in-render"
)


# =========================================================
# BASE DIRECTORY
# =========================================================

BASE_DIR = os.path.dirname(
    os.path.abspath(__file__)
)


# =========================================================
# RENDER PERSISTENT DISK
# =========================================================
#
# On Render:
#
# Persistent Disk mount path:
# /var/data
#
# You can also run locally without /var/data.
#

if os.path.exists("/var/data"):
    STORAGE_DIR = "/var/data"
else:
    STORAGE_DIR = os.path.join(
        BASE_DIR,
        "data"
    )


# =========================================================
# DATABASE / EXCEL
# =========================================================

DB_PATH = os.path.join(
    STORAGE_DIR,
    "vijay.db"
)

EXCEL_PATH = os.path.join(
    STORAGE_DIR,
    "VIJAY_Registration_Data.xlsx"
)


# =========================================================
# UPLOAD DIRECTORIES
# =========================================================

if os.path.exists("/var/data"):

    PHOTO_DIR = os.path.join(
        STORAGE_DIR,
        "photos"
    )

    CV_DIR = os.path.join(
        STORAGE_DIR,
        "cv"
    )

    QR_DIR = os.path.join(
        STORAGE_DIR,
        "qrcodes"
    )

else:

    PHOTO_DIR = os.path.join(
        BASE_DIR,
        "static",
        "uploads",
        "photos"
    )

    CV_DIR = os.path.join(
        BASE_DIR,
        "static",
        "uploads",
        "cv"
    )

    QR_DIR = os.path.join(
        BASE_DIR,
        "static",
        "qrcodes"
    )


# =========================================================
# UPLOAD SETTINGS
# =========================================================

ALLOWED_PHOTOS = {
    "jpg",
    "jpeg",
    "png",
    "webp"
}

ALLOWED_CV = {
    "pdf",
    "doc",
    "docx"
}


# Maximum complete request size
MAX_UPLOAD = 10 * 1024 * 1024

app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD


# =========================================================
# CREATE REQUIRED DIRECTORIES
# =========================================================

for folder in [
    STORAGE_DIR,
    PHOTO_DIR,
    CV_DIR,
    QR_DIR
]:

    os.makedirs(
        folder,
        exist_ok=True
    )


# =========================================================
# DATABASE CONNECTION
# =========================================================

def db():

    conn = sqlite3.connect(
        DB_PATH,
        timeout=30
    )

    conn.row_factory = sqlite3.Row

    return conn


# =========================================================
# DATABASE INITIALIZATION
# =========================================================

def init_db():

    conn = db()

    # -----------------------------------------------------
    # STUDENT REGISTRATIONS
    # -----------------------------------------------------

    conn.execute("""
        CREATE TABLE IF NOT EXISTS registrations (

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            registration_no TEXT UNIQUE NOT NULL,

            full_name TEXT NOT NULL,

            dob TEXT,

            gender TEXT,

            phone TEXT NOT NULL,

            email TEXT,

            parent_name TEXT,

            address TEXT,

            city TEXT,

            state TEXT,

            pincode TEXT,

            qualification TEXT,

            institution TEXT,

            course TEXT,

            academic_year TEXT,

            photo TEXT,

            pdf_file TEXT,

            qr_file TEXT,

            created_at TEXT NOT NULL
        )
    """)


    # -----------------------------------------------------
    # FACULTY
    # -----------------------------------------------------

    conn.execute("""
        CREATE TABLE IF NOT EXISTS faculty (

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            faculty_no TEXT UNIQUE NOT NULL,

            full_name TEXT NOT NULL,

            dob TEXT,

            gender TEXT,

            phone TEXT NOT NULL,

            email TEXT NOT NULL,

            address TEXT,

            city TEXT,

            state TEXT,

            pincode TEXT,

            highest_qualification TEXT,

            specialization TEXT,

            experience TEXT,

            previous_institution TEXT,

            designation TEXT,

            joining_date TEXT,

            expected_salary TEXT,

            cv_file TEXT,

            photo TEXT,

            created_at TEXT NOT NULL
        )
    """)


    conn.commit()

    conn.close()


# =========================================================
# FILE VALIDATION
# =========================================================

def allowed_file(
    filename,
    allowed
):

    return (
        "." in filename
        and
        filename.rsplit(
            ".",
            1
        )[1].lower() in allowed
    )


# =========================================================
# UNIQUE FILE NAME
# =========================================================

def unique_filename(filename):

    ext = filename.rsplit(
        ".",
        1
    )[1].lower()

    return (
        f"{uuid.uuid4().hex}.{ext}"
    )


# =========================================================
# SAVE UPLOAD
# =========================================================

def save_upload(
    file_obj,
    folder,
    allowed
):

    if not file_obj:
        return None

    if not file_obj.filename:
        return None

    if not allowed_file(
        file_obj.filename,
        allowed
    ):
        return None

    safe_name = secure_filename(
        file_obj.filename
    )

    filename = unique_filename(
        safe_name
    )

    file_path = os.path.join(
        folder,
        filename
    )

    file_obj.save(
        file_path
    )

    return filename


# =========================================================
# GENERATE NEXT REGISTRATION / FACULTY NUMBER
# =========================================================

def next_number(
    prefix,
    table,
    column
):

    conn = db()

    row = conn.execute(
        f"""
        SELECT {column}
        FROM {table}
        ORDER BY id DESC
        LIMIT 1
        """
    ).fetchone()

    conn.close()


    if not row or not row[column]:

        return f"{prefix}000001"


    m = re.search(
        r"(\d+)$",
        row[column]
    )


    number = (
        int(m.group(1)) + 1
        if m
        else 1
    )


    return f"{prefix}{number:06d}"


# =========================================================
# ADMIN LOGIN DECORATOR
# =========================================================

def admin_required(fn):

    @wraps(fn)
    def wrapper(
        *args,
        **kwargs
    ):

        if not session.get(
            "admin_logged_in"
        ):

            return redirect(
                url_for(
                    "admin_login",
                    next=request.path
                )
            )

        return fn(
            *args,
            **kwargs
        )

    return wrapper


# =========================================================
# EXCEL HEADER
# =========================================================

def excel_header(
    ws,
    headers
):

    for col, value in enumerate(
        headers,
        1
    ):

        cell = ws.cell(
            1,
            col,
            value
        )

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="0E2948"
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = ws.dimensions


# =========================================================
# EXCEL STYLING
# =========================================================

def style_excel(ws):

    thin = Side(
        style="thin",
        color="D9E2EC"
    )


    for row in ws.iter_rows():

        for cell in row:

            cell.border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )


    for col in range(
        1,
        ws.max_column + 1
    ):

        max_len = 0

        for row in range(
            1,
            min(
                ws.max_row,
                100
            ) + 1
        ):

            val = ws.cell(
                row,
                col
            ).value

            if val is not None:

                max_len = max(
                    max_len,
                    len(str(val))
                )


        ws.column_dimensions[
            get_column_letter(col)
        ].width = min(
            max(
                max_len + 2,
                12
            ),
            35
        )


# =========================================================
# UPDATE EXCEL
# =========================================================

def update_excel():

    conn = db()


    registrations = conn.execute(
        """
        SELECT *
        FROM registrations
        ORDER BY id DESC
        """
    ).fetchall()


    faculty = conn.execute(
        """
        SELECT *
        FROM faculty
        ORDER BY id DESC
        """
    ).fetchall()


    conn.close()


    wb = Workbook()


    # =====================================================
    # REGISTRATIONS SHEET
    # =====================================================

    ws = wb.active

    ws.title = "Registrations"


    reg_headers = [

        "ID",
        "Registration No",
        "Full Name",
        "DOB",
        "Gender",
        "Phone",
        "Email",
        "Parent / Guardian",
        "Address",
        "City",
        "State",
        "Pincode",
        "Qualification",
        "Institution",
        "Course",
        "Academic Year",
        "Photo File",
        "PDF File",
        "QR File",
        "Created At"
    ]


    excel_header(
        ws,
        reg_headers
    )


    for r in registrations:

        ws.append([

            r["id"],
            r["registration_no"],
            r["full_name"],
            r["dob"],
            r["gender"],
            r["phone"],
            r["email"],
            r["parent_name"],
            r["address"],
            r["city"],
            r["state"],
            r["pincode"],
            r["qualification"],
            r["institution"],
            r["course"],
            r["academic_year"],
            r["photo"],
            r["pdf_file"],
            r["qr_file"],
            r["created_at"]
        ])


    # =====================================================
    # FACULTY SHEET
    # =====================================================

    ws2 = wb.create_sheet(
        "Faculty"
    )


    faculty_headers = [

        "ID",
        "Faculty No",
        "Full Name",
        "DOB",
        "Gender",
        "Phone",
        "Email",
        "Address",
        "City",
        "State",
        "Pincode",
        "Highest Qualification",
        "Specialization",
        "Experience",
        "Previous Institution",
        "Designation",
        "Joining Date",
        "Expected Salary",
        "CV File",
        "Photo File",
        "Created At"
    ]


    excel_header(
        ws2,
        faculty_headers
    )


    for f in faculty:

        ws2.append([

            f["id"],
            f["faculty_no"],
            f["full_name"],
            f["dob"],
            f["gender"],
            f["phone"],
            f["email"],
            f["address"],
            f["city"],
            f["state"],
            f["pincode"],
            f["highest_qualification"],
            f["specialization"],
            f["experience"],
            f["previous_institution"],
            f["designation"],
            f["joining_date"],
            f["expected_salary"],
            f["cv_file"],
            f["photo"],
            f["created_at"]
        ])


    # =====================================================
    # INSTRUCTIONS
    # =====================================================

    ws3 = wb.create_sheet(
        "Instructions"
    )


    ws3["A1"] = (
        "VIJAY EDUCATIONAL SERVICES "
        "- DATA WORKBOOK"
    )


    ws3["A1"].font = Font(
        bold=True,
        size=16,
        color="0E2948"
    )


    ws3["A3"] = "Registrations"

    ws3["B3"] = (
        "Student/candidate registration records."
    )


    ws3["A4"] = "Faculty"

    ws3["B4"] = (
        "Faculty joining/application records."
    )


    ws3["A5"] = "Automatic update"

    ws3["B5"] = (
        "The workbook is regenerated automatically "
        "after each successful submission."
    )


    ws3["A6"] = "QR"

    ws3["B6"] = (
        "Scanning a registration QR opens the public "
        "registration page."
    )


    for sheet in [
        ws,
        ws2,
        ws3
    ]:

        style_excel(sheet)


    wb.save(
        EXCEL_PATH
    )


# =========================================================
# BUILD REGISTRATION PDF
# =========================================================

def build_registration_pdf(
    reg
):

    filename = (
        f"{reg['registration_no']}.pdf"
    )


    path = os.path.join(
        STORAGE_DIR,
        filename
    )


    doc = SimpleDocTemplate(

        path,

        pagesize=A4,

        rightMargin=14 * mm,

        leftMargin=14 * mm,

        topMargin=14 * mm,

        bottomMargin=14 * mm
    )


    styles = getSampleStyleSheet()


    title = ParagraphStyle(

        "TitleCustom",

        parent=styles["Title"],

        fontSize=20,

        textColor=colors.HexColor(
            "#0E2948"
        ),

        alignment=TA_CENTER,

        spaceAfter=4
    )


    subtitle = ParagraphStyle(

        "Sub",

        parent=styles["Normal"],

        fontSize=10,

        textColor=colors.HexColor(
            "#369F7B"
        ),

        alignment=TA_CENTER,

        spaceAfter=10
    )


    normal = ParagraphStyle(

        "NormalCustom",

        parent=styles["Normal"],

        fontSize=9.5,

        leading=13
    )


    section = ParagraphStyle(

        "Section",

        parent=styles["Heading2"],

        fontSize=12,

        textColor=colors.HexColor(
            "#0E2948"
        ),

        spaceBefore=8,

        spaceAfter=5
    )


    story = []


    # =====================================================
    # LOGO
    # =====================================================

    logo_path = os.path.join(

        BASE_DIR,

        "static",

        "logo.jpg"
    )


    if os.path.exists(
        logo_path
    ):

        logo = RLImage(

            logo_path,

            width=105 * mm,

            height=44 * mm
        )

        logo.hAlign = "CENTER"

        story.append(logo)


    story.append(

        Paragraph(

            "ONLINE EDUCATION PLATFORM "
            "FOR ENTRANCE & FOUNDATION",

            subtitle
        )
    )


    story.append(

        Paragraph(

            "REGISTRATION FORM",

            title
        )
    )


    story.append(

        Paragraph(

            f"<b>Registration No:</b> "
            f"{reg['registration_no']}",

            normal
        )
    )


    story.append(
        Spacer(1, 6)
    )


    # =====================================================
    # PHOTO
    # =====================================================

    if reg["photo"]:

        photo_path = os.path.join(

            PHOTO_DIR,

            reg["photo"]
        )


        if os.path.exists(
            photo_path
        ):

            p = RLImage(

                photo_path,

                width=30 * mm,

                height=36 * mm
            )

            p.hAlign = "RIGHT"

            story.append(p)

            story.append(
                Spacer(1, 4)
            )


    # =====================================================
    # PERSONAL DETAILS
    # =====================================================

    story.append(

        Paragraph(

            "PERSONAL DETAILS",

            section
        )
    )


    personal = [

        [
            "Full Name",
            reg["full_name"] or ""
        ],

        [
            "Date of Birth",
            reg["dob"] or ""
        ],

        [
            "Gender",
            reg["gender"] or ""
        ],

        [
            "Phone",
            reg["phone"] or ""
        ],

        [
            "Email",
            reg["email"] or ""
        ],

        [
            "Parent / Guardian",
            reg["parent_name"] or ""
        ],

        [
            "Address",
            reg["address"] or ""
        ],

        [
            "City / State / PIN",

            f"{reg['city'] or ''} / "
            f"{reg['state'] or ''} / "
            f"{reg['pincode'] or ''}"
        ]
    ]


    t = Table(

        personal,

        colWidths=[

            45 * mm,

            125 * mm
        ]
    )


    t.setStyle(

        TableStyle([

            (

                "GRID",

                (0, 0),

                (-1, -1),

                0.5,

                colors.HexColor(
                    "#D9E2EC"
                )
            ),

            (

                "BACKGROUND",

                (0, 0),

                (0, -1),

                colors.HexColor(
                    "#EAF7F2"
                )
            ),

            (

                "FONTNAME",

                (0, 0),

                (0, -1),

                "Helvetica-Bold"
            ),

            (

                "VALIGN",

                (0, 0),

                (-1, -1),

                "TOP"
            ),

            (

                "PADDING",

                (0, 0),

                (-1, -1),

                6
            )
        ])
    )


    story.append(t)


    # =====================================================
    # EDUCATION DETAILS
    # =====================================================

    story.append(

        Paragraph(

            "EDUCATION DETAILS",

            section
        )
    )


    education = [

        [
            "Qualification",
            reg["qualification"] or ""
        ],

        [
            "Institution",
            reg["institution"] or ""
        ],

        [
            "Course / Program",
            reg["course"] or ""
        ],

        [
            "Academic Year",
            reg["academic_year"] or ""
        ]
    ]


    t2 = Table(

        education,

        colWidths=[

            45 * mm,

            125 * mm
        ]
    )


    t2.setStyle(

        TableStyle([

            (

                "GRID",

                (0, 0),

                (-1, -1),

                0.5,

                colors.HexColor(
                    "#D9E2EC"
                )
            ),

            (

                "BACKGROUND",

                (0, 0),

                (0, -1),

                colors.HexColor(
                    "#EAF7F2"
                )
            ),

            (

                "FONTNAME",

                (0, 0),

                (0, -1),

                "Helvetica-Bold"
            ),

            (

                "VALIGN",

                (0, 0),

                (-1, -1),

                "TOP"
            ),

            (

                "PADDING",

                (0, 0),

                (-1, -1),

                6
            )
        ])
    )


    story.append(t2)


    story.append(
        Spacer(1, 12)
    )


    story.append(

        Paragraph(

            "I confirm that the information provided "
            "in this registration form is correct "
            "to the best of my knowledge.",

            normal
        )
    )


    story.append(
        Spacer(1, 22)
    )


    sign = Table(

        [

            [
                "Applicant Signature",
                "Authorized Signature"
            ],

            [
                "________________________",
                "________________________"
            ]
        ],

        colWidths=[

            85 * mm,

            85 * mm
        ]
    )


    sign.setStyle(

        TableStyle([

            (

                "ALIGN",

                (0, 0),

                (-1, -1),

                "CENTER"
            ),

            (

                "FONTNAME",

                (0, 0),

                (-1, 0),

                "Helvetica-Bold"
            ),

            (

                "TEXTCOLOR",

                (0, 0),

                (-1, 0),

                colors.HexColor(
                    "#0E2948"
                )
            ),

            (

                "TOPPADDING",

                (0, 0),

                (-1, -1),

                6
            )
        ])
    )


    story.append(sign)


    story.append(
        Spacer(1, 12)
    )


    story.append(

        Paragraph(

            f"Generated on: "
            f"{reg['created_at']}",

            normal
        )
    )


    doc.build(story)


    return filename


# =========================================================
# QR CODE
# =========================================================

def make_qr(
    registration_no
):

    public_url = url_for(

        "public_registration",

        registration_no=registration_no,

        _external=True
    )


    img = qrcode.make(
        public_url
    )


    filename = (
        f"{registration_no}.png"
    )


    img.save(

        os.path.join(

            QR_DIR,

            filename
        )
    )


    return filename


# =========================================================
# HOME
# =========================================================

@app.route("/")
def home():

    return render_template(
        "home.html"
    )


# =========================================================
# ABOUT
# =========================================================

@app.route("/about")
def about():

    return render_template(
        "about.html"
    )


# =========================================================
# ACADEMIC PLAN
# =========================================================

@app.route("/academic-plan")
def academic_plan():

    return render_template(
        "academic_plan.html"
    )


# =========================================================
# CAREERS
# =========================================================

@app.route("/careers")
def careers():

    return render_template(
        "careers.html"
    )


# =========================================================
# COURSES
# =========================================================

@app.route("/courses")
def courses():

    return render_template(
        "courses.html"
    )


# =========================================================
# STUDENT LOGIN
# =========================================================

@app.route(
    "/student-login",
    methods=["GET", "POST"]
)
def student_login():

    if request.method == "POST":

        email = request.form.get(
            "email",
            ""
        )

        password = request.form.get(
            "password",
            ""
        )

        # -------------------------------------------------
        # ADD YOUR STUDENT LOGIN VERIFICATION HERE
        # -------------------------------------------------

    return render_template(
        "student_login.html"
    )


# =========================================================
# CONTACT
# =========================================================

@app.route(
    "/contact",
    methods=["GET", "POST"]
)
def contact():

    if request.method == "POST":

        name = request.form.get(
            "name",
            ""
        ).strip()

        email = request.form.get(
            "email",
            ""
        ).strip()

        phone = request.form.get(
            "phone",
            ""
        ).strip()

        subject = request.form.get(
            "subject",
            ""
        ).strip()

        message = request.form.get(
            "message",
            ""
        ).strip()


        if not name or not email or not message:

            flash(
                "Please fill all required fields.",
                "error"
            )

            return redirect(
                url_for(
                    "contact"
                )
            )


        flash(
            "Your message has been submitted successfully.",
            "success"
        )


        return redirect(
            url_for(
                "contact"
            )
        )


    return render_template(
        "contact.html"
    )


# =========================================================
# STUDENT REGISTRATION
# =========================================================

@app.route(
    "/register",
    methods=["GET", "POST"]
)
def register():

    if request.method == "POST":

        required = [

            "full_name",

            "phone"
        ]


        if any(

            not request.form.get(
                x,
                ""
            ).strip()

            for x in required

        ):

            flash(
                "Please fill all required fields.",
                "error"
            )

            return render_template(
                "register.html"
            )


        # -------------------------------------------------
        # PHOTO
        # -------------------------------------------------

        photo_file = request.files.get(
            "photo"
        )


        photo = save_upload(

            photo_file,

            PHOTO_DIR,

            ALLOWED_PHOTOS
        )


        if photo_file and photo_file.filename and not photo:

            flash(
                "Photo must be JPG, JPEG, PNG or WEBP.",
                "error"
            )

            return render_template(
                "register.html"
            )


        # -------------------------------------------------
        # REGISTRATION NUMBER
        # -------------------------------------------------

        registration_no = next_number(

            "VJES",

            "registrations",

            "registration_no"
        )


        created_at = datetime.now().strftime(

            "%Y-%m-%d %H:%M:%S"
        )


        conn = db()


        try:

            conn.execute("""

                INSERT INTO registrations (

                    registration_no,

                    full_name,

                    dob,

                    gender,

                    phone,

                    email,

                    parent_name,

                    address,

                    city,

                    state,

                    pincode,

                    qualification,

                    institution,

                    course,

                    academic_year,

                    photo,

                    created_at

                )

                VALUES (

                    ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,

                    ?, ?, ?, ?, ?, ?, ?

                )

            """, (

                registration_no,

                request.form.get(
                    "full_name",
                    ""
                ).strip(),

                request.form.get(
                    "dob",
                    ""
                ),

                request.form.get(
                    "gender",
                    ""
                ),

                request.form.get(
                    "phone",
                    ""
                ).strip(),

                request.form.get(
                    "email",
                    ""
                ).strip(),

                request.form.get(
                    "parent_name",
                    ""
                ).strip(),

                request.form.get(
                    "address",
                    ""
                ).strip(),

                request.form.get(
                    "city",
                    ""
                ).strip(),

                request.form.get(
                    "state",
                    ""
                ).strip(),

                request.form.get(
                    "pincode",
                    ""
                ).strip(),

                request.form.get(
                    "qualification",
                    ""
                ).strip(),

                request.form.get(
                    "institution",
                    ""
                ).strip(),

                request.form.get(
                    "course",
                    ""
                ).strip(),

                request.form.get(
                    "academic_year",
                    ""
                ).strip(),

                photo,

                created_at
            ))


            conn.commit()


            row = conn.execute(

                """

                SELECT *

                FROM registrations

                WHERE registration_no = ?

                """,

                (registration_no,)

            ).fetchone()


            # -------------------------------------------------
            # CREATE PDF
            # -------------------------------------------------

            pdf_file = build_registration_pdf(
                row
            )


            # -------------------------------------------------
            # CREATE QR
            # -------------------------------------------------

            qr_file = make_qr(
                registration_no
            )


            # -------------------------------------------------
            # UPDATE DATABASE
            # -------------------------------------------------

            conn.execute(

                """

                UPDATE registrations

                SET pdf_file = ?,

                    qr_file = ?

                WHERE id = ?

                """,

                (

                    pdf_file,

                    qr_file,

                    row["id"]
                )
            )


            conn.commit()


        except Exception:

            conn.rollback()

            raise


        finally:

            conn.close()


        # -------------------------------------------------
        # UPDATE EXCEL
        # -------------------------------------------------

        update_excel()


        return redirect(

            url_for(

                "registration_success",

                registration_no=registration_no
            )
        )


    return render_template(
        "register.html"
    )


# =========================================================
# REGISTRATION SUCCESS
# =========================================================

@app.route(
    "/registration/success/<registration_no>"
)
def registration_success(
    registration_no
):

    conn = db()


    reg = conn.execute(

        """

        SELECT *

        FROM registrations

        WHERE registration_no = ?

        """,

        (registration_no,)

    ).fetchone()


    conn.close()


    if not reg:

        abort(404)


    return render_template(

        "success.html",

        reg=reg
    )


# =========================================================
# PUBLIC REGISTRATION
# =========================================================

@app.route(
    "/registration/<registration_no>"
)
def public_registration(
    registration_no
):

    conn = db()


    reg = conn.execute(

        """

        SELECT *

        FROM registrations

        WHERE registration_no = ?

        """,

        (registration_no,)

    ).fetchone()


    conn.close()


    if not reg:

        abort(404)


    return render_template(

        "public_registration.html",

        reg=reg
    )


# =========================================================
# REGISTRATION PDF
# =========================================================

@app.route(
    "/registration/<registration_no>/pdf"
)
def registration_pdf(
    registration_no
):

    conn = db()


    reg = conn.execute(

        """

        SELECT *

        FROM registrations

        WHERE registration_no = ?

        """,

        (registration_no,)

    ).fetchone()


    conn.close()


    if not reg or not reg["pdf_file"]:

        abort(404)


    path = os.path.join(

        STORAGE_DIR,

        reg["pdf_file"]
    )


    if not os.path.exists(path):

        abort(404)


    return send_file(

        path,

        as_attachment=True,

        download_name=reg["pdf_file"]
    )


# =========================================================
# PHOTO FILE
# =========================================================

@app.route(
    "/uploads/photos/<filename>"
)
def uploaded_photo(
    filename
):

    return send_from_directory(

        PHOTO_DIR,

        filename
    )


# =========================================================
# QR FILE
# =========================================================

@app.route(
    "/qrcodes/<filename>"
)
def qr_code(
    filename
):

    return send_from_directory(

        QR_DIR,

        filename
    )


# =========================================================
# FACULTY
# =========================================================

@app.route(
    "/faculty",
    methods=["GET", "POST"]
)
def faculty():

    if request.method == "POST":

        fields = [

            "full_name",

            "phone",

            "email"
        ]


        if any(

            not request.form.get(
                x,
                ""
            ).strip()

            for x in fields

        ):

            flash(
                "Please fill all required faculty fields.",
                "error"
            )

            return render_template(
                "faculty.html"
            )


        # -------------------------------------------------
        # PHOTO
        # -------------------------------------------------

        photo_file = request.files.get(
            "photo"
        )


        photo = save_upload(

            photo_file,

            PHOTO_DIR,

            ALLOWED_PHOTOS
        )


        # -------------------------------------------------
        # CV
        # -------------------------------------------------

        cv_file = request.files.get(
            "cv"
        )


        cv = save_upload(

            cv_file,

            CV_DIR,

            ALLOWED_CV
        )


        if photo_file and photo_file.filename and not photo:

            flash(
                "Photo must be JPG, JPEG, PNG or WEBP.",
                "error"
            )

            return render_template(
                "faculty.html"
            )


        if cv_file and cv_file.filename and not cv:

            flash(
                "CV must be PDF, DOC or DOCX.",
                "error"
            )

            return render_template(
                "faculty.html"
            )


        # -------------------------------------------------
        # FACULTY NUMBER
        # -------------------------------------------------

        faculty_no = next_number(

            "FAC",

            "faculty",

            "faculty_no"
        )


        created_at = datetime.now().strftime(

            "%Y-%m-%d %H:%M:%S"
        )


        conn = db()


        try:

            conn.execute("""

                INSERT INTO faculty (

                    faculty_no,

                    full_name,

                    dob,

                    gender,

                    phone,

                    email,

                    address,

                    city,

                    state,

                    pincode,

                    highest_qualification,

                    specialization,

                    experience,

                    previous_institution,

                    designation,

                    joining_date,

                    expected_salary,

                    cv_file,

                    photo,

                    created_at

                )

                VALUES (

                    ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,

                    ?, ?, ?, ?, ?, ?, ?, ?, ?, ?

                )

            """, (

                faculty_no,

                request.form.get(
                    "full_name",
                    ""
                ).strip(),

                request.form.get(
                    "dob",
                    ""
                ),

                request.form.get(
                    "gender",
                    ""
                ),

                request.form.get(
                    "phone",
                    ""
                ).strip(),

                request.form.get(
                    "email",
                    ""
                ).strip(),

                request.form.get(
                    "address",
                    ""
                ).strip(),

                request.form.get(
                    "city",
                    ""
                ).strip(),

                request.form.get(
                    "state",
                    ""
                ).strip(),

                request.form.get(
                    "pincode",
                    ""
                ).strip(),

                request.form.get(
                    "highest_qualification",
                    ""
                ).strip(),

                request.form.get(
                    "specialization",
                    ""
                ).strip(),

                request.form.get(
                    "experience",
                    ""
                ).strip(),

                request.form.get(
                    "previous_institution",
                    ""
                ).strip(),

                request.form.get(
                    "designation",
                    ""
                ).strip(),

                request.form.get(
                    "joining_date",
                    ""
                ),

                request.form.get(
                    "expected_salary",
                    ""
                ).strip(),

                cv,

                photo,

                created_at
            ))


            conn.commit()


        except Exception:

            conn.rollback()

            raise


        finally:

            conn.close()


        update_excel()


        return render_template(

            "faculty_success.html",

            faculty_no=faculty_no
        )


    return render_template(
        "faculty.html"
    )


# =========================================================
# ADMIN LOGIN
# =========================================================

@app.route(
    "/admin/login",
    methods=["GET", "POST"]
)
def admin_login():

    if request.method == "POST":

        username = request.form.get(
            "username",
            ""
        )


        password = request.form.get(
            "password",
            ""
        )


        admin_user = os.environ.get(

            "ADMIN_USERNAME",

            "admin"
        )


        admin_pass = os.environ.get(

            "ADMIN_PASSWORD",

            "admin123"
        )


        if (

            username == admin_user

            and

            password == admin_pass

        ):

            session[
                "admin_logged_in"
            ] = True


            next_page = request.args.get(
                "next"
            )


            if next_page:

                return redirect(
                    next_page
                )


            return redirect(

                url_for(
                    "admin_dashboard"
                )
            )


        flash(

            "Invalid admin username or password.",

            "error"
        )


    return render_template(
        "admin_login.html"
    )


# =========================================================
# ADMIN LOGOUT
# =========================================================

@app.route(
    "/admin/logout"
)
def admin_logout():

    session.clear()


    return redirect(
        url_for("home")
    )


# =========================================================
# ADMIN DASHBOARD
# =========================================================

@app.route(
    "/admin"
)
@admin_required
def admin_dashboard():

    conn = db()


    registrations = conn.execute(

        """

        SELECT *

        FROM registrations

        ORDER BY id DESC

        """

    ).fetchall()


    faculty_rows = conn.execute(

        """

        SELECT *

        FROM faculty

        ORDER BY id DESC

        """

    ).fetchall()


    conn.close()


    return render_template(

        "admin.html",

        registrations=registrations,

        faculty=faculty_rows
    )


# =========================================================
# ADMIN EXCEL DOWNLOAD
# =========================================================

@app.route(
    "/admin/excel"
)
@admin_required
def download_excel():

    update_excel()


    if not os.path.exists(
        EXCEL_PATH
    ):

        abort(404)


    return send_file(

        EXCEL_PATH,

        as_attachment=True,

        download_name=(
            "VIJAY_Registration_Data.xlsx"
        )
    )


# =========================================================
# ADMIN REGISTRATION PDF
# =========================================================

@app.route(
    "/admin/registration/<registration_no>/pdf"
)
@admin_required
def admin_registration_pdf(
    registration_no
):

    return registration_pdf(
        registration_no
    )


# =========================================================
# ADMIN CV DOWNLOAD
# =========================================================

@app.route(
    "/admin/cv/<filename>"
)
@admin_required
def admin_cv(
    filename
):

    return send_from_directory(

        CV_DIR,

        filename,

        as_attachment=True
    )


# =========================================================
# FILE SIZE ERROR
# =========================================================

@app.errorhandler(413)
def too_large(_):

    flash(

        "Upload is too large. "
        "Maximum total request size is 10 MB.",

        "error"
    )


    return redirect(

        request.referrer

        or

        url_for("home")
    )


# =========================================================
# GENERAL ERROR HANDLER
# =========================================================

@app.errorhandler(500)
def server_error(error):

    return render_template(
        "500.html"
    ), 500


# =========================================================
# STARTUP
# =========================================================

# Initialize database when Gunicorn starts.
init_db()


# Create Excel workbook if it does not exist.
if not os.path.exists(
    EXCEL_PATH
):

    try:

        update_excel()

    except Exception as e:

        print(
            "Excel initialization error:",
            e
        )


# =========================================================
# LOCAL DEVELOPMENT
# =========================================================

if __name__ == "__main__":

    port = int(
        os.environ.get(
            "PORT",
            5000
        )
    )


    app.run(

        host="0.0.0.0",

        port=port,

        debug=os.environ.get(
            "FLASK_DEBUG",
            "0"
        ) == "1"
    )

